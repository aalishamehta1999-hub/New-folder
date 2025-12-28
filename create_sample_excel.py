import openpyxl

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Wedding Contacts"

# Add header row
ws['A1'] = 'Name'
ws['B1'] = 'Phone'
ws['C1'] = 'Mehendi'
ws['D1'] = 'Sangeet'
ws['E1'] = 'Wedding'
ws['F1'] = 'Reception'

# Add sample data with Yes/No
data = [
    ['Ganesh Kumar', '+919876543210', 'Yes', 'Yes', 'Yes', 'Yes'],
    ['Priya Sharma', '+918765432109', 'No', 'Yes', 'Yes', 'Yes'],
    ['Rahul Verma', '+917654321098', 'Yes', 'No', 'Yes', 'Yes'],
    ['Anjali Patel', '+916543210987', 'Yes', 'Yes', 'No', 'Yes'],
    ['Vikram Singh', '+915432109876', 'Yes', 'Yes', 'Yes', 'No'],
]

for idx, row_data in enumerate(data, start=2):
    ws[f'A{idx}'] = row_data[0]
    ws[f'B{idx}'] = row_data[1]
    ws[f'C{idx}'] = row_data[2]
    ws[f'D{idx}'] = row_data[3]
    ws[f'E{idx}'] = row_data[4]
    ws[f'F{idx}'] = row_data[5]

# Save the file
wb.save('sample_contacts.xlsx')
print("✅ Sample Excel file created: sample_contacts.xlsx")
print("\nFormat:")
print("- Name | Phone | Mehendi | Sangeet | Wedding | Reception")
print("- Use 'Yes' or 'No' for each category")
print("\nYou can now upload this file to test the application!")
